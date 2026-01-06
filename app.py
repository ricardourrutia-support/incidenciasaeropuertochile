import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime, date, timedelta
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule


APP_VERSION = "v2026-01-05_FULL_KPIS_ONE_DETAIL_CF_FIXED"

st.set_page_config(page_title="Ausentismo e Incidencias Operativas", layout="wide")
st.title("Plataforma de Gestión de Ausentismo e Incidencias Operativas")
st.caption(APP_VERSION)
st.sidebar.success(f"APP RUNNING: {APP_VERSION}")

# =========================
# Estilo (Cabify-ish)
# =========================
CABIFY_HEADER = "362065"
CABIFY_ACCENT = "E83C96"
GRID = "D9D9D9"
WHITE = "FFFFFF"

thin = Side(style="thin", color=GRID)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

FILL_GRAY = PatternFill("solid", fgColor="EDEDED")   # gris suave
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")   # amarillo suave
FILL_HEADER = PatternFill("solid", fgColor=CABIFY_HEADER)

def style_header_row(ws, row=1, fill_hex=CABIFY_HEADER):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(color=WHITE, bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize_columns(ws, max_width=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for c in col[:600]:
            v = c.value
            if v is None:
                continue
            s = str(v)
            max_len = max(max_len, len(s))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

# =========================
# Opciones (dropdowns)
# =========================
TIPO_OPTS = ["Inasistencia", "Marcaje/Turno", "No Procede"]
CLASIF_OPTS = ["Seleccionar", "Injustificada", "Permiso", "Licencia", "Vacaciones", "Compensado", "No Procede"]

# =========================
# Helpers parsing
# =========================
def normalize_rut(x):
    if pd.isna(x):
        return ""
    return str(x).strip().upper().replace(".", "").replace(" ", "")

def try_parse_date_any(x):
    if pd.isna(x):
        return pd.NaT
    return pd.to_datetime(x, errors="coerce", dayfirst=True)

def _clean_cell(x):
    if pd.isna(x):
        return ""
    return str(x).replace("\u00A0", " ").strip()

def _norm_colname(s):
    s = "" if s is None else str(s)
    s = s.replace("\u00A0", " ").lower().strip()
    for ch in [" ", ".", "-", "_", "\n", "\t", "\r", ":", ";", ",", "(", ")"]:
        s = s.replace(ch, "")
    return s

def find_col(df, candidates):
    norm = {_norm_colname(c): c for c in df.columns}
    for c in candidates:
        k = _norm_colname(c)
        if k in norm:
            return norm[k]
    return None

def read_raw(file, sheet=0):
    name = getattr(file, "name", "").lower()
    if name.endswith(".xls"):
        return pd.read_excel(file, sheet_name=sheet, header=None, engine="xlrd")
    return pd.read_excel(file, sheet_name=sheet, header=None, engine="openpyxl")

def ffill_row(values):
    out, last = [], ""
    for v in values:
        s = _clean_cell(v)
        if s:
            last = s
            out.append(s)
        else:
            out.append(last)
    return out

def norm_recinto(x):
    if pd.isna(x):
        return "Sin Marca"
    s = str(x).strip()
    su = s.upper().replace("SÍ", "SI")
    if su in ["SI", "S"]:
        return "Sí"
    if su in ["NO", "N"]:
        return "No"
    if su in ["SIN COORDENADAS", "SINCOORDENADAS"]:
        return "No"
    if su == "":
        return "Sin Marca"
    return s

def safe_time_str(x):
    if pd.isna(x):
        return ""
    try:
        t = pd.to_datetime(x, errors="coerce")
        if pd.isna(t):
            s = str(x).strip()
            return s[:5] if len(s) >= 5 else s
        tt = t.time()
        return f"{tt.hour:02d}:{tt.minute:02d}"
    except Exception:
        s = str(x).strip()
        return s[:5] if len(s) >= 5 else s

# =========================
# Horario parsing (para Horas_Plan)
# =========================
TIME_RE = re.compile(r"(\d{1,2}):(\d{2})(?::(\d{2}))?")

def normalize_horario_str(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    if s == "" or s.lower() in ["nan", "none"]:
        return ""
    matches = TIME_RE.findall(s)
    if len(matches) < 2:
        return s
    (h1, m1, _), (h2, m2, _) = matches[0], matches[1]
    return f"{int(h1):02d}:{m1}-{int(h2):02d}:{m2}"

def parse_horario_to_hours(horario: str) -> float:
    hs = normalize_horario_str(horario)
    if not isinstance(hs, str) or "-" not in hs:
        return np.nan
    a, b = hs.split("-", 1)
    try:
        t1 = pd.to_datetime(a).time()
        t2 = pd.to_datetime(b).time()
    except Exception:
        return np.nan
    dt1 = datetime(2000, 1, 1, t1.hour, t1.minute)
    dt2 = datetime(2000, 1, 1, t2.hour, t2.minute)
    if dt2 <= dt1:
        dt2 += timedelta(days=1)
    return (dt2 - dt1).total_seconds() / 3600.0

# =========================
# Lecturas específicas
# =========================
def read_asistencia(file):
    raw = read_raw(file)
    if len(raw) < 3:
        raise RuntimeError("Asistencia: necesito 2 filas header + datos.")
    h1 = ffill_row(raw.iloc[0].tolist())
    h2 = [_clean_cell(v) for v in raw.iloc[1].tolist()]

    start = 1 if (h1 and h1[0] == "" and h2 and h2[0] == "") else 0
    h1, h2 = h1[start:], h2[start:]

    cols = []
    for g, s in zip(h1, h2):
        gn, sn = _norm_colname(g), _norm_colname(s)
        if gn == "entrada" and sn == "fecha":
            cols.append("Fecha Entrada")
        elif gn == "entrada" and sn == "hora":
            cols.append("Hora Entrada")
        elif gn == "salida" and sn == "fecha":
            cols.append("Fecha Salida")
        elif gn == "salida" and sn == "hora":
            cols.append("Hora Salida")
        else:
            if s:
                if g and _norm_colname(g) != _norm_colname(s):
                    cols.append(f"{g} {s}".strip())
                else:
                    cols.append(s)
            else:
                cols.append(g)

    cols = [_clean_cell(c) if c else f"COL_{i}" for i, c in enumerate(cols)]
    df = raw.iloc[2:, start:].copy()
    df.columns = cols
    df = df.dropna(how="all")
    return df, raw

def read_inasist(file):
    raw = read_raw(file)
    header = None
    for i in range(min(160, len(raw))):
        row = [_norm_colname(v) for v in raw.iloc[i].tolist()]
        if "rut" in row and ("dia" in row or "día" in row):
            header = i
            break
    if header is None:
        raise RuntimeError("Inasistencias: no pude detectar encabezado (RUT + Día).")
    cols = [_clean_cell(c) for c in raw.iloc[header].tolist()]
    df = raw.iloc[header + 1 :].copy()
    df.columns = cols
    df = df.dropna(how="all")
    return df, raw

# =========================
# Sidebar - inputs
# =========================
with st.sidebar:
    st.header("Inputs")
    f_asist = st.file_uploader("1) Asistencia (XLS/XLSX)", type=["xls", "xlsx"])
    f_inas = st.file_uploader("2) Inasistencias (XLS/XLSX)", type=["xls", "xlsx"])
    f_plan = st.file_uploader("3) Planificación de Turnos (CSV)", type=["csv"])
    f_cod = st.file_uploader("4) Codificación de Turno (CSV)", type=["csv"])

    st.divider()
    st.subheader("Filtros / reglas")
    only_area = st.text_input("Filtrar Área (opcional)", value="AEROPUERTO")
    umbral_diff_h = st.number_input("Diferencia mínima (horas) para marcar incidencia", value=0.5, step=0.25, min_value=0.0)

if not all([f_asist, f_inas, f_plan, f_cod]):
    st.info("Sube los 4 archivos para comenzar.")
    st.stop()

# =========================
# Load
# =========================
df_asist, _ = read_asistencia(f_asist)
df_inas, _ = read_inasist(f_inas)
df_plan = pd.read_csv(f_plan)
df_cod = pd.read_csv(f_cod)

# =========================
# Normalizaciones + filtros base
# =========================
c_rut_a = find_col(df_asist, ["RUT"])
if not c_rut_a:
    st.error("Asistencia: no encontré columna RUT.")
    st.stop()

c_rut_i = find_col(df_inas, ["RUT"])
c_dia_i = find_col(df_inas, ["Día", "Dia"])
if not c_rut_i or not c_dia_i:
    st.error("Inasistencias: faltan columnas RUT y/o Día.")
    st.stop()

if "Fecha Entrada" not in df_asist.columns:
    st.error("Asistencia: falta 'Fecha Entrada'.")
    st.stop()

df_asist["RUT_norm"] = df_asist[c_rut_a].apply(normalize_rut)
df_asist["Fecha_base"] = df_asist["Fecha Entrada"].apply(try_parse_date_any).dt.date
df_asist["HoraEntrada_str"] = df_asist["Hora Entrada"].apply(safe_time_str) if "Hora Entrada" in df_asist.columns else ""
df_asist["HoraSalida_str"] = df_asist["Hora Salida"].apply(safe_time_str) if "Hora Salida" in df_asist.columns else ""

c_rec_in = find_col(df_asist, ["Dentro de Recinto(Entrada)", "Dentro del Recinto (Entrada)"])
c_rec_out = find_col(df_asist, ["Dentro de Recinto(Salida)", "Dentro del Recinto (Salida)"])

if not c_rec_in:
    df_asist["Dentro de Recinto(Entrada)"] = "Sin Marca"
    c_rec_in = "Dentro de Recinto(Entrada)"
if not c_rec_out:
    df_asist["Dentro de Recinto(Salida)"] = "Sin Marca"
    c_rec_out = "Dentro de Recinto(Salida)"

df_asist[c_rec_in] = df_asist[c_rec_in].apply(norm_recinto)
df_asist[c_rec_out] = df_asist[c_rec_out].apply(norm_recinto)

df_inas["RUT_norm"] = df_inas[c_rut_i].apply(normalize_rut)
df_inas["Fecha_base"] = df_inas[c_dia_i].apply(try_parse_date_any).dt.date

def maybe_filter_area(df, col="Área"):
    if only_area and col in df.columns:
        return df[df[col].astype(str).str.upper().str.contains(only_area.upper(), na=False)].copy()
    return df

df_asist = maybe_filter_area(df_asist, "Área")
df_inas = maybe_filter_area(df_inas, "Área")
df_plan = maybe_filter_area(df_plan, "Área")

# =========================
# Filtro previo por Especialidad
# =========================
c_esp_a = find_col(df_asist, ["Especialidad"])
c_esp_i = find_col(df_inas, ["Especialidad"])

roster = []
if c_esp_a:
    roster.append(df_asist[["RUT_norm", c_esp_a]].rename(columns={c_esp_a: "Especialidad"}))
if c_esp_i:
    roster.append(df_inas[["RUT_norm", c_esp_i]].rename(columns={c_esp_i: "Especialidad"}))

if roster:
    roster_df = pd.concat(roster, ignore_index=True).dropna(subset=["RUT_norm"]).drop_duplicates()
    roster_df["Especialidad"] = roster_df["Especialidad"].astype(str).str.strip()
    esp_list = sorted([e for e in roster_df["Especialidad"].unique().tolist() if e and e.lower() != "nan"])
else:
    roster_df = pd.DataFrame(columns=["RUT_norm", "Especialidad"])
    esp_list = []

with st.sidebar:
    if esp_list:
        esp_sel = st.multiselect("Filtrar por Especialidad (opcional)", options=esp_list, default=esp_list)
    else:
        esp_sel = []

if esp_sel:
    allowed_ruts = set(roster_df[roster_df["Especialidad"].isin(esp_sel)]["RUT_norm"].tolist())
    df_asist = df_asist[df_asist["RUT_norm"].isin(allowed_ruts)].copy()
    df_inas = df_inas[df_inas["RUT_norm"].isin(allowed_ruts)].copy()

# =========================
# Selector fechas
# =========================
fixed_plan_cols = ["Nombre del Colaborador", "RUT", "Área", "Supervisor"]
plan_date_cols = [c for c in df_plan.columns if c not in fixed_plan_cols]

def try_parse_plan_col_to_date(c):
    dt = pd.to_datetime(str(c), errors="coerce", dayfirst=True)
    return dt.date() if not pd.isna(dt) else None

plan_dates = [try_parse_plan_col_to_date(c) for c in plan_date_cols]
plan_dates = [d for d in plan_dates if d is not None]

all_dates = []
all_dates += [d for d in df_asist["Fecha_base"].dropna().tolist()]
all_dates += [d for d in df_inas["Fecha_base"].dropna().tolist()]
all_dates += plan_dates

if not all_dates:
    st.error("No pude inferir fechas para el selector de periodo.")
    st.stop()

min_d, max_d = min(all_dates), max(all_dates)
c1, c2 = st.columns(2)
with c1:
    date_from = st.date_input("Desde", value=min_d, min_value=min_d, max_value=max_d)
with c2:
    date_to = st.date_input("Hasta", value=max_d, min_value=min_d, max_value=max_d)

if date_from > date_to:
    st.error("Rango inválido.")
    st.stop()

df_asist = df_asist[(df_asist["Fecha_base"] >= date_from) & (df_asist["Fecha_base"] <= date_to)].copy()
df_inas = df_inas[(df_inas["Fecha_base"] >= date_from) & (df_inas["Fecha_base"] <= date_to)].copy()

# =========================
# Planificación wide -> long (sin L y vacíos)
# =========================
if "RUT" not in df_plan.columns:
    st.error("Planificación CSV: falta columna RUT.")
    st.stop()

df_plan["RUT_norm"] = df_plan["RUT"].apply(normalize_rut)

date_cols = [c for c in df_plan.columns if c not in fixed_plan_cols]
plan_long = df_plan.melt(
    id_vars=[c for c in fixed_plan_cols if c in df_plan.columns],
    value_vars=date_cols,
    var_name="Fecha_col",
    value_name="Turno_Cod"
)

plan_long["Fecha"] = plan_long["Fecha_col"].apply(lambda x: try_parse_plan_col_to_date(x))
plan_long = plan_long.dropna(subset=["Fecha"]).copy()
plan_long["RUT_norm"] = plan_long["RUT"].apply(normalize_rut)
plan_long["Turno_Cod"] = plan_long["Turno_Cod"].astype(str).str.strip()
plan_long.loc[plan_long["Turno_Cod"].isin(["", "nan", "None", "-"]), "Turno_Cod"] = ""

plan_long = plan_long[(plan_long["Fecha"] >= date_from) & (plan_long["Fecha"] <= date_to)].copy()
plan_long = plan_long[plan_long["Turno_Cod"] != ""].copy()
plan_long = plan_long[~plan_long["Turno_Cod"].astype(str).str.upper().isin(["L"])].copy()

if esp_sel:
    plan_long = plan_long[plan_long["RUT_norm"].isin(allowed_ruts)].copy()

# =========================
# Codificación: solo para Horario Planificado / Horas_Plan
# =========================
c_sigla = find_col(df_cod, ["Sigla", "SIGLA"])
c_hor = find_col(df_cod, ["Horario", "HORARIO"])
if not c_sigla or not c_hor:
    st.error("Codificación: no encontré columnas Sigla y Horario.")
    st.stop()

cod_map = df_cod[[c_sigla, c_hor]].copy()
cod_map.columns = ["Turno_Cod", "Horario_raw"]
cod_map["Turno_Cod"] = cod_map["Turno_Cod"].astype(str).str.strip()
cod_map["Horario Planificado"] = cod_map["Horario_raw"].apply(normalize_horario_str)
cod_map["Horas_Plan"] = cod_map["Horario Planificado"].apply(parse_horario_to_hours)

plan_long = plan_long.merge(
    cod_map[["Turno_Cod", "Horario Planificado", "Horas_Plan"]],
    on="Turno_Cod",
    how="left"
)

# =========================
# Construcción Detalle
# =========================
def combine_dt(fecha_col, hora_col):
    if pd.isna(fecha_col):
        return pd.NaT
    d = pd.to_datetime(fecha_col, errors="coerce", dayfirst=True)
    if pd.isna(d):
        return pd.NaT
    if hora_col is None or str(hora_col).strip() == "":
        return pd.Timestamp(d)
    t = pd.to_datetime(hora_col, errors="coerce")
    if pd.isna(t):
        return pd.Timestamp(d)
    tt = t.time()
    return datetime(d.year, d.month, d.day, tt.hour, tt.minute)

def worked_hours(dt_in, dt_out):
    if pd.isna(dt_in) or pd.isna(dt_out):
        return np.nan
    delta = (dt_out - dt_in).total_seconds() / 3600.0
    if delta < 0:
        delta += 24.0
    return delta

df_asist["dt_in"] = [combine_dt(f, h) for f, h in zip(df_asist.get("Fecha Entrada"), df_asist.get("Hora Entrada", [""] * len(df_asist)))]
df_asist["dt_out"] = [combine_dt(f, h) for f, h in zip(df_asist.get("Fecha Salida", df_asist.get("Fecha Entrada")), df_asist.get("Hora Salida", [""] * len(df_asist)))]
df_asist["Horas_Trab"] = [worked_hours(a, b) for a, b in zip(df_asist["dt_in"], df_asist["dt_out"])]

asist_merge = df_asist.merge(
    plan_long[["RUT_norm", "Fecha", "Horario Planificado", "Horas_Plan"]],
    left_on=["RUT_norm", "Fecha_base"],
    right_on=["RUT_norm", "Fecha"],
    how="left"
)

asist_merge["Marcas_Fuera"] = (
    (asist_merge[c_rec_in].astype(str).str.strip().str.upper() == "NO") |
    (asist_merge[c_rec_out].astype(str).str.strip().str.upper() == "NO")
).astype(int)

asist_merge["Diff_h"] = (asist_merge["Horas_Plan"].astype(float) - asist_merge["Horas_Trab"].astype(float))

mask_inc = (
    (~asist_merge["Horas_Plan"].isna()) &
    (
        (asist_merge["Diff_h"].fillna(0) >= float(umbral_diff_h)) |
        (asist_merge["Marcas_Fuera"] == 1)
    )
)
df_inc = asist_merge[mask_inc].copy()

c_cod_a = find_col(df_asist, ["Código", "Codigo"])
c_nom = find_col(df_asist, ["Nombre"])
c_ap1 = find_col(df_asist, ["Primer Apellido", "PrimerApellido"])
c_ap2 = find_col(df_asist, ["Segundo Apellido", "SegundoApellido"])
c_esp = find_col(df_asist, ["Especialidad"])
c_sup = find_col(df_asist, ["Supervisor"])

def col_or_blank(df, c):
    return df[c] if c and c in df.columns else ""

def fmt1(x):
    try:
        if pd.isna(x):
            return ""
        return f"{float(x):.1f}"
    except Exception:
        return ""

df_inc_detalle = pd.DataFrame({
    "Fecha": df_inc["Fecha_base"],
    "Código": col_or_blank(df_inc, c_cod_a),
    "RUT": df_inc[c_rut_a],
    "Nombre": col_or_blank(df_inc, c_nom),
    "Primer Apellido": col_or_blank(df_inc, c_ap1),
    "Segundo Apellido": col_or_blank(df_inc, c_ap2),
    "Especialidad": col_or_blank(df_inc, c_esp),
    "Supervisor": col_or_blank(df_inc, c_sup),
    "Turno Marcado": df_inc["HoraEntrada_str"].astype(str) + "-" + df_inc["HoraSalida_str"].astype(str),
    "Horario Planificado": df_inc["Horario Planificado"],
    "Dentro de Recinto(Entrada)": df_inc[c_rec_in],
    "Dentro de Recinto(Salida)": df_inc[c_rec_out],
    "Tipo_Incidencia": "Marcaje/Turno",
    "Detalle": (
        "HorasPlan=" + df_inc["Horas_Plan"].map(fmt1) +
        " | HorasTrab=" + df_inc["Horas_Trab"].map(fmt1) +
        " | Diff_h=" + df_inc["Diff_h"].map(fmt1) +
        " | MarcasFuera=" + df_inc["Marcas_Fuera"].astype(str)
    ),
    "Clasificación Manual": "Seleccionar",
    "Minutos Retraso": 0,
    "Minutos Salida Anticipada": 0,
    "RUT_norm": df_inc["RUT_norm"],
})

c_cod_i = find_col(df_inas, ["Código", "Codigo"])
c_nom_i = find_col(df_inas, ["Nombre"])
c_ap1_i = find_col(df_inas, ["Primer Apellido", "PrimerApellido"])
c_ap2_i = find_col(df_inas, ["Segundo Apellido", "SegundoApellido"])
c_esp_i = find_col(df_inas, ["Especialidad"])
c_sup_i = find_col(df_inas, ["Supervisor"])
c_mot = find_col(df_inas, ["Motivo"])

df_inas_detalle = pd.DataFrame({
    "Fecha": df_inas["Fecha_base"],
    "Código": col_or_blank(df_inas, c_cod_i),
    "RUT": df_inas[c_rut_i],
    "Nombre": col_or_blank(df_inas, c_nom_i),
    "Primer Apellido": col_or_blank(df_inas, c_ap1_i),
    "Segundo Apellido": col_or_blank(df_inas, c_ap2_i),
    "Especialidad": col_or_blank(df_inas, c_esp_i),
    "Supervisor": col_or_blank(df_inas, c_sup_i),
    "Turno Marcado": "",
    "Horario Planificado": "",
    "Dentro de Recinto(Entrada)": "Sin Marca",
    "Dentro de Recinto(Salida)": "Sin Marca",
    "Tipo_Incidencia": "Inasistencia",
    "Detalle": "Motivo=" + (col_or_blank(df_inas, c_mot).astype(str) if c_mot else ""),
    "Clasificación Manual": "Seleccionar",
    "Minutos Retraso": 0,
    "Minutos Salida Anticipada": 0,
    "RUT_norm": df_inas["RUT_norm"],
})

# Auto por motivo
if c_mot and c_mot in df_inas.columns:
    mot = df_inas[c_mot].astype(str).str.strip().str.upper()
    df_inas_detalle.loc[mot == "P", "Clasificación Manual"] = "Permiso"
    df_inas_detalle.loc[mot == "L", "Clasificación Manual"] = "Licencia"
    df_inas_detalle.loc[mot == "V", "Clasificación Manual"] = "Vacaciones"
    df_inas_detalle.loc[mot == "C", "Clasificación Manual"] = "Compensado"

detalle = pd.concat([df_inc_detalle, df_inas_detalle], ignore_index=True)
detalle["Fecha"] = pd.to_datetime(detalle["Fecha"], errors="coerce").dt.date
detalle = detalle.sort_values(["Fecha", "RUT"]).reset_index(drop=True)

# =========================
# UI: Detalle editable
# =========================
st.subheader("Detalle (editable en la app)")
edited = st.data_editor(
    detalle.drop(columns=["RUT_norm"], errors="ignore"),
    use_container_width=True,
    num_rows="dynamic",
    column_config={
        "Fecha": st.column_config.DateColumn(format="DD/MM/YYYY"),
        "Tipo_Incidencia": st.column_config.SelectboxColumn(options=TIPO_OPTS),
        "Clasificación Manual": st.column_config.SelectboxColumn(options=CLASIF_OPTS),
        "Minutos Retraso": st.column_config.NumberColumn(min_value=0, step=1),
        "Minutos Salida Anticipada": st.column_config.NumberColumn(min_value=0, step=1),
    }
)

# =========================
# Validación obligatoria (bloquea descarga)
# =========================
def to_int_safe(x):
    try:
        if pd.isna(x) or str(x).strip() == "":
            return 0
        return int(float(x))
    except Exception:
        return 0

edited_valid = edited.copy()
edited_valid["Minutos Retraso"] = edited_valid["Minutos Retraso"].apply(to_int_safe)
edited_valid["Minutos Salida Anticipada"] = edited_valid["Minutos Salida Anticipada"].apply(to_int_safe)
edited_valid["Minutos_Total"] = edited_valid["Minutos Retraso"] + edited_valid["Minutos Salida Anticipada"]

mask_need_minutes = (
    (edited_valid["Tipo_Incidencia"] == "Marcaje/Turno") &
    (edited_valid["Clasificación Manual"] == "Injustificada") &
    (edited_valid["Minutos_Total"] <= 0)
)
mask_minutes_not_allowed = (
    (edited_valid["Tipo_Incidencia"] != "Marcaje/Turno") &
    (edited_valid["Minutos_Total"] > 0)
)

invalid = edited_valid[mask_need_minutes | mask_minutes_not_allowed].copy()
if len(invalid) > 0:
    st.error("Hay registros inválidos. Para descargar el Excel debes corregirlos.")
    st.stop()

edited = edited_valid.drop(columns=["Minutos_Total"], errors="ignore")

# =========================
# Excel helpers
# =========================
def write_df(ws, df: pd.DataFrame):
    ws.append(list(df.columns))
    for r in df.itertuples(index=False):
        ws.append(list(r))

def add_dropdown(ws, col_letter: str, start_row: int, end_row: int, options: list, prompt: str):
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=False)
    dv.showDropDown = False  # flecha visible
    dv.promptTitle = "Seleccionar"
    dv.prompt = prompt
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

def protect_sheet(ws):
    ws.protection.sheet = True
    ws.protection.enable()

def lock_all_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

def unlock_range(ws, col_letter, start_row, end_row):
    for r in range(start_row, end_row + 1):
        ws[f"{col_letter}{r}"].protection = Protection(locked=False)

def add_minutes_conditional_formatting(ws, col_tipo, col_clas, col_mr, col_ms, start_row, end_row):
    rng = f"{col_mr}{start_row}:{col_ms}{end_row}"

    # Amarillo primero (falta minutos cuando corresponde) y detiene
    warn_formula = f'=AND(${col_tipo}{start_row}="Marcaje/Turno",${col_clas}{start_row}="Injustificada",(${col_mr}{start_row}+${col_ms}{start_row})=0)'
    rule_warn = FormulaRule(formula=[warn_formula], fill=FILL_WARN, stopIfTrue=True)
    ws.conditional_formatting.add(rng, rule_warn)

    # Gris: cuando NO corresponde editar minutos
    gray_formula = f'=NOT(AND(${col_tipo}{start_row}="Marcaje/Turno",${col_clas}{start_row}="Injustificada"))'
    rule_gray = FormulaRule(formula=[gray_formula], fill=FILL_GRAY)
    ws.conditional_formatting.add(rng, rule_gray)

def add_minutes_validation(ws, col_tipo, col_clas, col_mr, col_ms, start_row, end_row):
    msg = (
        "Reglas:\n"
        "- Si Tipo=Marcaje/Turno y Clasificación=Injustificada => MinR+MinS > 0\n"
        "- Si Tipo != Marcaje/Turno => MinR y MinS deben ser 0"
    )
    for r in range(start_row, end_row + 1):
        formula_ok = (
            f'=IF(AND(${col_tipo}{r}="Marcaje/Turno",${col_clas}{r}="Injustificada"),'
            f'(${col_mr}{r}+${col_ms}{r})>0,'
            f'(${col_mr}{r}+${col_ms}{r})=0)'
        )
        dv_mr = DataValidation(type="custom", formula1=formula_ok, allow_blank=True)
        dv_mr.errorTitle = "Minutos inválidos"
        dv_mr.error = msg
        ws.add_data_validation(dv_mr)
        dv_mr.add(f"{col_mr}{r}")

        dv_ms = DataValidation(type="custom", formula1=formula_ok, allow_blank=True)
        dv_ms.errorTitle = "Minutos inválidos"
        dv_ms.error = msg
        ws.add_data_validation(dv_ms)
        dv_ms.add(f"{col_ms}{r}")

# =========================
# Excel builder (incluye KPIs, Cumplimiento, Planificacion_long)
# =========================
def build_excel(edited_df: pd.DataFrame, plan_long_df: pd.DataFrame, date_from: date, date_to: date) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    # Listas (oculta)
    ws_list = wb.create_sheet("Listas")
    ws_list["A1"] = "Tipo_Incidencia"
    for i, v in enumerate(TIPO_OPTS, start=2):
        ws_list[f"A{i}"] = v
    ws_list["C1"] = "Clasificación Manual"
    for i, v in enumerate(CLASIF_OPTS, start=2):
        ws_list[f"C{i}"] = v
    style_header_row(ws_list, 1, CABIFY_HEADER)
    ws_list.sheet_state = "hidden"

    # -------------------------
    # Detalle (HOJA PRINCIPAL)
    # -------------------------
    ws_det = wb.create_sheet("Detalle")

    # Encabezado filtro supervisor (solo ayuda + dropdown)
    ws_det["A1"] = "Filtrar Supervisor (usa el filtro de la tabla):"
    ws_det["A1"].font = Font(bold=True, color=CABIFY_ACCENT)
    ws_det["B1"] = ""
    ws_det["A2"] = "Tip: Activa el filtro del encabezado y filtra la columna Supervisor."
    ws_det["A2"].font = Font(italic=True)

    det_df = edited_df.copy()
    det_df["Fecha"] = pd.to_datetime(det_df["Fecha"], errors="coerce").dt.date

    start_table_row = 4
    ws_det.append([])  # row 3 vacío
    ws_det.append(list(det_df.columns))  # row 4 headers

    for r in det_df.itertuples(index=False):
        ws_det.append(list(r))

    # style headers row 4
    style_header_row(ws_det, start_table_row, CABIFY_HEADER)

    n_rows = len(det_df) + start_table_row
    cols = list(det_df.columns)

    def col_letter_of(name):
        return get_column_letter(cols.index(name) + 1)

    # AutoFilter sobre la tabla (Excel + Sheets friendly)
    last_col = get_column_letter(len(cols))
    ws_det.auto_filter.ref = f"A{start_table_row}:{last_col}{n_rows}"
    ws_det.freeze_panes = f"A{start_table_row+1}"

    # Dropdown supervisor en B1 (si existe columna Supervisor)
    if "Supervisor" in det_df.columns:
        sup_values = sorted(
            [s for s in det_df["Supervisor"].dropna().astype(str).str.strip().unique().tolist()
             if s and s.lower() != "nan"]
        )
    else:
        sup_values = []

    ws_list["E1"] = "Supervisores"
    for i, v in enumerate(sup_values, start=2):
        ws_list[f"E{i}"] = v

    if sup_values:
        end_row = 1 + len(sup_values)
        dv_sup = DataValidation(type="list", formula1=f"=Listas!$E$2:$E${end_row}", allow_blank=False)
        dv_sup.showDropDown = False
        ws_det.add_data_validation(dv_sup)
        dv_sup.add("B1")
        ws_det["B1"].value = sup_values[0]

    # Dropdowns sobre tabla
    if "Tipo_Incidencia" in cols:
        add_dropdown(ws_det, col_letter_of("Tipo_Incidencia"), start_table_row+1, n_rows, TIPO_OPTS,
                     "Inasistencia / Marcaje/Turno / No Procede")
    if "Clasificación Manual" in cols:
        add_dropdown(ws_det, col_letter_of("Clasificación Manual"), start_table_row+1, n_rows, CLASIF_OPTS,
                     "Selecciona clasificación")

    # Fecha format
    if "Fecha" in cols:
        c_fecha = cols.index("Fecha") + 1
        for rr in range(start_table_row+1, n_rows+1):
            ws_det.cell(rr, c_fecha).number_format = "DD/MM/YYYY"

    # Validación + formato condicional en minutos
    if all(c in cols for c in ["Tipo_Incidencia", "Clasificación Manual", "Minutos Retraso", "Minutos Salida Anticipada"]):
        col_tipo = col_letter_of("Tipo_Incidencia")
        col_clas = col_letter_of("Clasificación Manual")
        col_mr = col_letter_of("Minutos Retraso")
        col_ms = col_letter_of("Minutos Salida Anticipada")

        add_minutes_validation(ws_det, col_tipo, col_clas, col_mr, col_ms, start_table_row+1, n_rows)
        add_minutes_conditional_formatting(ws_det, col_tipo, col_clas, col_mr, col_ms, start_table_row+1, n_rows)

    autosize_columns(ws_det)

    # Protección: bloquear todo excepto inputs y B1
    lock_all_cells(ws_det)
    ws_det["B1"].protection = Protection(locked=False)

    for nm in ["Tipo_Incidencia", "Clasificación Manual", "Minutos Retraso", "Minutos Salida Anticipada"]:
        if nm in cols:
            unlock_range(ws_det, col_letter_of(nm), start_table_row+1, n_rows)

    protect_sheet(ws_det)

    # -------------------------
    # Planificacion_long (con fórmulas dependientes de Detalle)
    # -------------------------
    ws_plan = wb.create_sheet("Planificacion_long")
    pl = plan_long_df.copy()

    needed = ["Fecha", "RUT", "Nombre del Colaborador", "Área", "Supervisor", "Turno_Cod", "Horario Planificado", "Horas_Plan"]
    for n in needed:
        if n not in pl.columns:
            pl[n] = ""
    pl = pl[needed].copy()
    pl["Fecha"] = pd.to_datetime(pl["Fecha"], errors="coerce").dt.date

    pl["Ausente_Injustificada"] = ""
    pl["Min_Retraso_Injust"] = ""
    pl["Min_Salida_Injust"] = ""
    pl["Horas_Incid_Injust"] = ""

    write_df(ws_plan, pl)
    style_header_row(ws_plan, 1, CABIFY_HEADER)
    autosize_columns(ws_plan)
    ws_plan.freeze_panes = "A2"

    pl_cols = list(pl.columns)
    L_pl_fecha = get_column_letter(pl_cols.index("Fecha") + 1)
    L_pl_rut = get_column_letter(pl_cols.index("RUT") + 1)
    L_pl_hplan = get_column_letter(pl_cols.index("Horas_Plan") + 1)
    L_pl_aus = get_column_letter(pl_cols.index("Ausente_Injustificada") + 1)
    L_pl_mr = get_column_letter(pl_cols.index("Min_Retraso_Injust") + 1)
    L_pl_ms = get_column_letter(pl_cols.index("Min_Salida_Injust") + 1)
    L_pl_hinc = get_column_letter(pl_cols.index("Horas_Incid_Injust") + 1)

    det_cols = list(det_df.columns)
    # Detalle sheet starts header row at start_table_row
    # ranges start at start_table_row+1 for data
    det_data_start = start_table_row+1

    def dcol(name): return get_column_letter(det_cols.index(name) + 1)

    L_det_fecha = dcol("Fecha")
    L_det_rut = dcol("RUT")
    L_det_tipo = dcol("Tipo_Incidencia")
    L_det_clas = dcol("Clasificación Manual")
    L_det_mr = dcol("Minutos Retraso")
    L_det_ms = dcol("Minutos Salida Anticipada")

    # Use whole-column refs (works in Excel/Sheets) but OK
    # Note: Detalle headers at row 4 doesn't matter for SUMIFS ranges; we use full columns.
    for r in range(2, len(pl) + 2):
        ws_plan[f"{L_pl_aus}{r}"].value = (
            f'=IF(COUNTIFS(Detalle!${L_det_fecha}:${L_det_fecha},{L_pl_fecha}{r},'
            f'Detalle!${L_det_rut}:${L_det_rut},{L_pl_rut}{r},'
            f'Detalle!${L_det_tipo}:${L_det_tipo},"Inasistencia",'
            f'Detalle!${L_det_clas}:${L_det_clas},"Injustificada")>0,1,0)'
        )
        ws_plan[f"{L_pl_mr}{r}"].value = (
            f'=SUMIFS(Detalle!${L_det_mr}:${L_det_mr},'
            f'Detalle!${L_det_fecha}:${L_det_fecha},{L_pl_fecha}{r},'
            f'Detalle!${L_det_rut}:${L_det_rut},{L_pl_rut}{r},'
            f'Detalle!${L_det_tipo}:${L_det_tipo},"Marcaje/Turno",'
            f'Detalle!${L_det_clas}:${L_det_clas},"Injustificada")'
        )
        ws_plan[f"{L_pl_ms}{r}"].value = (
            f'=SUMIFS(Detalle!${L_det_ms}:${L_det_ms},'
            f'Detalle!${L_det_fecha}:${L_det_fecha},{L_pl_fecha}{r},'
            f'Detalle!${L_det_rut}:${L_det_rut},{L_pl_rut}{r},'
            f'Detalle!${L_det_tipo}:${L_det_tipo},"Marcaje/Turno",'
            f'Detalle!${L_det_clas}:${L_det_clas},"Injustificada")'
        )
        ws_plan[f"{L_pl_hinc}{r}"].value = f'=({L_pl_mr}{r}+{L_pl_ms}{r})/60'

    lock_all_cells(ws_plan)
    protect_sheet(ws_plan)

    # -------------------------
    # Cumplimiento
    # -------------------------
    ws_c = wb.create_sheet("Cumplimiento")
    base = pl[["RUT", "Nombre del Colaborador", "Supervisor", "Área"]].drop_duplicates().copy()
    base = base.sort_values(["Nombre del Colaborador", "RUT"]).reset_index(drop=True)

    out = base.copy()
    out["Turnos_Planificados"] = ""
    out["Inasistencias_Injustificadas"] = ""
    out["Min_Retraso_Injust"] = ""
    out["Min_Salida_Injust"] = ""
    out["Horas_Incid_Injust"] = ""
    out["Cumplimiento_%"] = ""
    out["Horas_programadas"] = ""
    out["Horas_perdidas_injust"] = ""
    out["Ausentismo_%"] = ""

    write_df(ws_c, out)
    style_header_row(ws_c, 1, CABIFY_HEADER)
    autosize_columns(ws_c)
    ws_c.freeze_panes = "A2"

    c_cols = list(out.columns)
    L_c_rut = get_column_letter(c_cols.index("RUT") + 1)
    L_c_tp = get_column_letter(c_cols.index("Turnos_Planificados") + 1)
    L_c_ina = get_column_letter(c_cols.index("Inasistencias_Injustificadas") + 1)
    L_c_mr = get_column_letter(c_cols.index("Min_Retraso_Injust") + 1)
    L_c_ms = get_column_letter(c_cols.index("Min_Salida_Injust") + 1)
    L_c_hinc = get_column_letter(c_cols.index("Horas_Incid_Injust") + 1)
    L_c_cump = get_column_letter(c_cols.index("Cumplimiento_%") + 1)
    L_c_hprog = get_column_letter(c_cols.index("Horas_programadas") + 1)
    L_c_hperd = get_column_letter(c_cols.index("Horas_perdidas_injust") + 1)
    L_c_ausp = get_column_letter(c_cols.index("Ausentismo_%") + 1)

    for r in range(2, len(out) + 2):
        ws_c[f"{L_c_tp}{r}"].value = f'=COUNTIF(Planificacion_long!${L_pl_rut}:${L_pl_rut},{L_c_rut}{r})'
        ws_c[f"{L_c_ina}{r}"].value = f'=SUMIF(Planificacion_long!${L_pl_rut}:${L_pl_rut},{L_c_rut}{r},Planificacion_long!${L_pl_aus}:${L_pl_aus})'
        ws_c[f"{L_c_mr}{r}"].value = f'=SUMIF(Planificacion_long!${L_pl_rut}:${L_pl_rut},{L_c_rut}{r},Planificacion_long!${L_pl_mr}:${L_pl_mr})'
        ws_c[f"{L_c_ms}{r}"].value = f'=SUMIF(Planificacion_long!${L_pl_rut}:${L_pl_rut},{L_c_rut}{r},Planificacion_long!${L_pl_ms}:${L_pl_ms})'
        ws_c[f"{L_c_hinc}{r}"].value = f'=SUMIF(Planificacion_long!${L_pl_rut}:${L_pl_rut},{L_c_rut}{r},Planificacion_long!${L_pl_hinc}:${L_pl_hinc})'

        # Cumplimiento: (turnos injustificados / turnos planificados)
        ws_c[f"{L_c_cump}{r}"].value = (
            f'=IF({L_c_tp}{r}=0,"",MAX(0,1-(({L_c_ina}{r}+'
            f'SUMPRODUCT((Planificacion_long!${L_pl_rut}:${L_pl_rut}={L_c_rut}{r})*(Planificacion_long!${L_pl_hinc}:${L_pl_hinc}>0))'
            f')/{L_c_tp}{r})))'
        )
        ws_c[f"{L_c_cump}{r}"].number_format = "0.00%"

        ws_c[f"{L_c_hprog}{r}"].value = f'=SUMIF(Planificacion_long!${L_pl_rut}:${L_pl_rut},{L_c_rut}{r},Planificacion_long!${L_pl_hplan}:${L_pl_hplan})'
        ws_c[f"{L_c_hperd}{r}"].value = (
            f'=SUMIFS(Planificacion_long!${L_pl_hplan}:${L_pl_hplan},'
            f'Planificacion_long!${L_pl_rut}:${L_pl_rut},{L_c_rut}{r},'
            f'Planificacion_long!${L_pl_aus}:${L_pl_aus},1)'
            f'+{L_c_hinc}{r}'
        )
        ws_c[f"{L_c_ausp}{r}"].value = f'=IF({L_c_hprog}{r}=0,"",{L_c_hperd}{r}/{L_c_hprog}{r})'
        ws_c[f"{L_c_ausp}{r}"].number_format = "0.00%"

    lock_all_cells(ws_c)
    protect_sheet(ws_c)

    # -------------------------
    # KPIs diarios (matriz)
    # -------------------------
    ws_k = wb.create_sheet("KPIs_diarios")
    fechas = pd.date_range(date_from, date_to, freq="D").date.tolist()

    kpis = [
        "Turnos_planificados",
        "Ausencias_injustificadas",
        "Min_Retraso_Injust",
        "Min_Salida_Injust",
        "Horas_Incid_Injust",
        "Cumplimiento_%",
        "Horas_programadas",
        "Horas_perdidas_injust",
        "Ausentismo_%",
    ]

    ws_k.append(["KPI"] + fechas)
    for k in kpis:
        ws_k.append([k] + [""] * len(fechas))

    style_header_row(ws_k, 1, CABIFY_HEADER)
    autosize_columns(ws_k)
    ws_k.freeze_panes = "B2"

    # referencias por columna
    for j, _ in enumerate(fechas, start=2):
        colL = get_column_letter(j)
        head = f"{colL}1"

        ws_k[f"{colL}2"].value = f'=COUNTIF(Planificacion_long!${L_pl_fecha}:${L_pl_fecha},{head})'
        ws_k[f"{colL}3"].value = f'=SUMIF(Planificacion_long!${L_pl_fecha}:${L_pl_fecha},{head},Planificacion_long!${L_pl_aus}:${L_pl_aus})'
        ws_k[f"{colL}4"].value = f'=SUMIF(Planificacion_long!${L_pl_fecha}:${L_pl_fecha},{head},Planificacion_long!${L_pl_mr}:${L_pl_mr})'
        ws_k[f"{colL}5"].value = f'=SUMIF(Planificacion_long!${L_pl_fecha}:${L_pl_fecha},{head},Planificacion_long!${L_pl_ms}:${L_pl_ms})'
        ws_k[f"{colL}6"].value = f'=SUMIF(Planificacion_long!${L_pl_fecha}:${L_pl_fecha},{head},Planificacion_long!${L_pl_hinc}:${L_pl_hinc})'
        ws_k[f"{colL}7"].value = (
            f'=IF({colL}2=0,"",MAX(0,1-(({colL}3+'
            f'SUMPRODUCT((Planificacion_long!${L_pl_fecha}:${L_pl_fecha}={head})*(Planificacion_long!${L_pl_hinc}:${L_pl_hinc}>0))'
            f')/{colL}2)))'
        )
        ws_k[f"{colL}7"].number_format = "0.00%"
        ws_k[f"{colL}8"].value = f'=SUMIF(Planificacion_long!${L_pl_fecha}:${L_pl_fecha},{head},Planificacion_long!${L_pl_hplan}:${L_pl_hplan})'
        ws_k[f"{colL}9"].value = (
            f'=SUMIFS(Planificacion_long!${L_pl_hplan}:${L_pl_hplan},'
            f'Planificacion_long!${L_pl_fecha}:${L_pl_fecha},{head},'
            f'Planificacion_long!${L_pl_aus}:${L_pl_aus},1)'
            f'+{colL}6'
        )
        ws_k[f"{colL}10"].value = f'=IF({colL}8=0,"",{colL}9/{colL}8)'
        ws_k[f"{colL}10"].number_format = "0.00%"

    lock_all_cells(ws_k)
    protect_sheet(ws_k)

    # proteger listas también
    lock_all_cells(ws_list)
    protect_sheet(ws_list)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

excel_bytes = build_excel(edited, plan_long, date_from, date_to)

st.subheader("Descarga")
st.download_button(
    "Descargar Excel (Detalle + KPIs + Validaciones + Formato condicional)",
    data=excel_bytes,
    file_name=f"reporte_ausentismo_incidencias_{date_from}_{date_to}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
